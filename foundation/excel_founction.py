import xlrd
import openpyxl
import openpyxl.styles
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_value_from_excel(file_path, sheet_name, value_index):
    """
    查找指定单元格的值
    :param file_path: excel文件路径
    :param sheet_name: sheet索引
    :param value_index: 单元格索引
    :return: 返回的值
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell_value = sheet[value_index].value
    return cell_value


def write_save_value_to_excel(file_path, sheet_name, value_index, value):
    """
    往指定单元格写值
    :param file_path: excel文件路径
    :param sheet_name: sheet名称
    :param value_index: 单元格索引
    :param value: 需要写入的值
    :return: null
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet[value_index] = value
    workbook.save(file_path)
    return 0


def cell_fill_color(file_path, sheet_name, cell_index, color):
    """
    填充指定单元格颜色
    :param file_path: excel文件路径
    :param sheet_name: sheet名称
    :param cell_index: 单元格索引
    :param color: 需要填充的颜色
    :return:
    """
    fill_color = PatternFill("solid", fgColor="FFFFFF")
    if color == 'red':
        fill_color = PatternFill("solid", fgColor="FF0000")
    elif color == 'green':
        fill_color = PatternFill("solid", fgColor="00FF00")
    elif color == 'blue':
        fill_color = PatternFill("solid", fgColor="0000FF")
    elif color == 'white':
        fill_color = PatternFill("solid", fgColor="FFFFFF")
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet[cell_index].fill = fill_color
    workbook.save(file_path)
    return 0

